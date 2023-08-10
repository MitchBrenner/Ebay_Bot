# class for scrapping information from ebay listings
from selenium.webdriver.remote.webelement import WebElement
from selenium.webdriver.common.by import By
from ebay_scraper.cosine_similarity import CosineSimilarity as similarity
import openpyxl as xl


class EbayReport:
    def __init__(self, el: WebElement):
        self.listings_box = el

    def get_listings(self):
        listings = self.listings_box.find_elements(
            By.CSS_SELECTOR, 'li[class="s-item s-item__pl-on-bottom"]'
        )
        listings_titles = []
        for listing in listings:
            title = (
                listing.find_element(By.CSS_SELECTOR, 'span[role="heading"]')
                .get_attribute("innerHTML")
                .strip()[12:-9]
            )
            listings_titles.append(title)
        return listings_titles

    def create_unique_listings(listings: list):
        # instatiate unique listing dict
        unique_listings = dict()
        # add first element form listings to unique_listings
        unique_listings[listings[0]] = 1
        # modify listings so that you don't look at the first element again
        listings = listings[1:]

        # iterate through all the listings
        found_match = False
        for listing in listings:
            most_similar_val = 0
            most_similar_el = None
            # compare listing to each unique listing
            for key, val in unique_listings.items():
                # if listings are similar incr val
                is_similar, sim_val = similarity.is_similar(listing, key)
                if is_similar:
                    found_match = True
                    if sim_val > most_similar_val:
                        most_similar_val = sim_val
                        most_similar_el = key
            # if match found update val of listing
            if found_match:
                unique_listings[most_similar_el] += 1
            # if match not found add new listing to unique listings
            else:
                unique_listings[listing] = 1

            found_match = False

        sorted_listings = dict(
            sorted(unique_listings.items(), key=lambda item: item[1], reverse=True)
        )

        return sorted_listings

    def create_sheet(data: list, keywords):
        keywords = keywords.lower().replace(" ", "_")

        workbook = xl.Workbook()
        sheet = workbook.active

        # Set fixed width for the first column
        sheet.column_dimensions["A"].width = 70

        # Insert keys in the first column
        for row, key in enumerate(data.keys(), start=1):
            sheet.cell(row=row, column=1, value=key)

        # Insert values in the second column
        for row, value in enumerate(data.values(), start=1):
            sheet.cell(row=row, column=2, value=value)

        workbook.save(f"{keywords}_report.xlsx")
